from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from db_config import get_connection
from datetime import datetime
from tareas_routes import tareas_bp, crear_tarjeta_trello  # Importar el blueprint y la función de Trello
import hashlib
import bcrypt
from collections import Counter
import re
from flask_login import login_required, current_user, LoginManager, UserMixin, login_user, logout_user
from werkzeug.security import check_password_hash, generate_password_hash
import openpyxl
import io
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from helpers import rol_requerido
from flask_socketio import SocketIO, emit
import requests
import logging
from config import TRELLO_API_KEY, TRELLO_API_TOKEN
from models import Notification
from notificaciones_utils import crear_y_emitir_notificacion
from werkzeug.utils import secure_filename
import os
from PIL import Image
from PIL import UnidentifiedImageError
from helpers import equipo_requerido
import shutil
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape as rl_landscape
from reportlab.lib.utils import ImageReader

app = Flask(__name__)
app.secret_key = 'tu_clave_secreta_aqui'  # Necesario para los mensajes flash

# Registrar el blueprint de tareas
app.register_blueprint(tareas_bp)

# Inicializar LoginManager
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Inicializar SocketIO
socketio = SocketIO(app, async_mode='threading')

# Modelo de usuario compatible con Flask-Login
class Usuario(UserMixin):
    pass

@login_manager.user_loader
def load_user(user_id):
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT * FROM empleados WHERE id = %s", (user_id,))
    usuario = cursor.fetchone()
    conexion.close()
    if usuario:
        user = Usuario()
        user.id = usuario['id']
        user.nombre = usuario['nombre']
        user.apellido = usuario['apellido']
        user.email = usuario['correo']
        user.rol = usuario['rol']
        user.telefono = usuario.get('telefono', '')
        user.password = usuario['password']
        user.equipo_id = usuario.get('equipo_id')
        # Puedes agregar más campos si los necesitas
        return user
    return None

@app.route('/')
@login_required
def index():
    conexion = get_connection()
    cursor = conexion.cursor()
    cursor.execute("SELECT * FROM empleados WHERE rol IN ('lider', 'practicante') ORDER BY nombre")
    empleados = cursor.fetchall()
    conexion.close()
    return render_template('index.html', empleados=empleados)

@app.route('/buscar', methods=['GET'])
def buscar():
    query = request.args.get('q', '')
    rol_filtro = request.args.get('rol', '')
    conexion = get_connection()
    cursor = conexion.cursor()
    sql = "SELECT * FROM empleados WHERE rol IN ('lider', 'practicante')"
    params = []
    if query:
        sql += " AND (nombre LIKE %s OR apellido LIKE %s OR correo LIKE %s)"
        params.extend([f'%{query}%', f'%{query}%', f'%{query}%'])
    if rol_filtro:
        sql += " AND rol = %s"
        params.append(rol_filtro)
    sql += " ORDER BY nombre"
    cursor.execute(sql, tuple(params))
    empleados = cursor.fetchall()
    conexion.close()
    return render_template('index.html', empleados=empleados, query=query, rol_filtro=rol_filtro)

@app.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar(id):
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT * FROM equipos ORDER BY nombre")
    equipos = cursor.fetchall()
    roles = ['jefe', 'lider', 'practicante']

    if request.method == 'POST':
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        correo = request.form['correo']
        telefono = request.form.get('telefono', '')
        equipo_id = request.form.get('equipo_id')
        rol = request.form.get('rol')
        habilidades = request.form.get('habilidades', '')
        fortalezas = request.form.get('fortalezas', '')
        horas_disponibles = request.form.get('horas_disponibles', 40)

        # Obtener equipo anterior
        cursor.execute("SELECT equipo_id, correo FROM empleados WHERE id=%s", (id,))
        row = cursor.fetchone()
        equipo_anterior = row['equipo_id']
        correo_anterior = row['correo']

        try:
            cursor.execute("""
                UPDATE empleados 
                SET nombre=%s, apellido=%s, correo=%s, telefono=%s, equipo_id=%s, rol=%s, habilidades=%s, horas_disponibles=%s
                WHERE id=%s
            """, (nombre, apellido, correo, telefono, equipo_id, rol, habilidades, horas_disponibles, id))
            # Si cambió de equipo y era líder/practicante
            if rol in ['lider', 'practicante'] and equipo_anterior and equipo_anterior != int(equipo_id or 0):
                # Obtener idBoard del equipo anterior
                cursor.execute("SELECT idBoard FROM equipos WHERE id=%s", (equipo_anterior,))
                row = cursor.fetchone()
                idBoard_anterior = row['idBoard'] if row else None
                if idBoard_anterior:
                    eliminar_de_tablero_trello(idBoard_anterior, correo_anterior)
                # Dejar tareas pendientes sin asignar
                cursor.execute("""
                    UPDATE tareas_trello SET empleado_id = NULL
                    WHERE empleado_id = %s AND estado != 'completada'
                """, (id,))
            # Actualizar fortalezas
            cursor.execute("DELETE FROM empleado_fortalezas WHERE empleado_id=%s", (id,))
            if rol in ['lider', 'practicante']:
                if habilidades:
                    habilidades_list = [h.strip() for h in habilidades.split(',') if h.strip()]
                    for habilidad in habilidades_list:
                        cursor.execute("SELECT id FROM habilidades WHERE nombre=%s", (habilidad,))
                        hab_row = cursor.fetchone()
                        if hab_row:
                            habilidad_id = hab_row['id']
                        else:
                            cursor.execute("INSERT INTO habilidades (nombre) VALUES (%s)", (habilidad,))
                            habilidad_id = cursor.lastrowid
                        cursor.execute("INSERT IGNORE INTO empleado_habilidades (empleado_id, habilidad_id) VALUES (%s, %s)", (id, habilidad_id))
                # Guardar fortalezas en la tabla relacional si se ingresan
                if fortalezas:
                    fortalezas_list = [f.strip() for f in fortalezas.split(',') if f.strip()]
                    for fortaleza in fortalezas_list:
                        cursor.execute("SELECT id FROM fortalezas WHERE nombre=%s", (fortaleza,))
                        fort_row = cursor.fetchone()
                        if fort_row:
                            fortaleza_id = fort_row['id']
                        else:
                            cursor.execute("INSERT INTO fortalezas (nombre) VALUES (%s)", (fortaleza,))
                            fortaleza_id = cursor.lastrowid
                        cursor.execute("INSERT IGNORE INTO empleado_fortalezas (empleado_id, fortaleza_id) VALUES (%s, %s)", (id, fortaleza_id))
            conexion.commit()
            flash('Empleado actualizado exitosamente!', 'success')
        except Exception as e:
            flash(f'Error al actualizar empleado: {str(e)}', 'danger')
        finally:
            conexion.close()
        return redirect(url_for('index'))

    cursor.execute("SELECT * FROM empleados WHERE id=%s", (id,))
    empleado = cursor.fetchone()
    # Obtener fortalezas actuales del empleado
    cursor.execute("""
        SELECT f.nombre FROM fortalezas f
        JOIN empleado_fortalezas ef ON f.id = ef.fortaleza_id
        WHERE ef.empleado_id = %s
    """, (id,))
    fortalezas_actuales = ', '.join([row['nombre'] for row in cursor.fetchall()])
    conexion.close()
    return render_template('editar.html', empleado=empleado, equipos=equipos, roles=roles, fortalezas_actuales=fortalezas_actuales)

@app.route('/eliminar/<int:id>')
def eliminar(id):
    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)
        # Obtener equipo y correo antes de eliminar
        cursor.execute("SELECT equipo_id, correo FROM empleados WHERE id=%s", (id,))
        row = cursor.fetchone()
        equipo_id = row['equipo_id'] if row else None
        correo = row['correo'] if row else None
        # Eliminar de Trello si corresponde
        if equipo_id and correo:
            cursor.execute("SELECT idBoard FROM equipos WHERE id=%s", (equipo_id,))
            row = cursor.fetchone()
            idBoard = row['idBoard'] if row else None
            if idBoard:
                eliminar_de_tablero_trello(idBoard, correo)
        # Dejar tareas pendientes sin asignar
        cursor.execute("UPDATE tareas_trello SET empleado_id = NULL WHERE empleado_id = %s AND estado != 'completada'", (id,))
        cursor.execute("DELETE FROM empleados_trello WHERE empleado_id = %s", (id,))
        cursor.execute("DELETE FROM empleados WHERE id=%s", (id,))
        conexion.commit()
        flash('Empleado eliminado exitosamente!', 'success')
    except Exception as e:
        flash(f'Error al eliminar empleado: {str(e)}', 'danger')
    finally:
        conexion.close()
    return redirect(url_for('index'))

@app.route('/reiniciar', methods=['GET', 'POST'])
@login_required
def reiniciar():
    # Solo permitir acceso a jefes
    if current_user.rol != 'jefe':
        flash('No tienes permiso para acceder a esta función.', 'danger')
        return redirect(url_for('index'))
    if request.method == 'POST':
        password = request.form.get('password')
        if not password:
            flash('Debes ingresar la contraseña especial.', 'danger')
            return render_template('reiniciar_confirmar.html')
        conexion = get_connection()
        cursor = conexion.cursor()
        cursor.execute("SELECT valor FROM configuracion WHERE clave = 'password_reinicio'")
        row = cursor.fetchone()
        if not row:
            flash('No se ha configurado la contraseña especial.', 'danger')
            conexion.close()
            return render_template('reiniciar_confirmar.html')
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        if password_hash != row[0]:
            flash('Contraseña incorrecta.', 'danger')
            conexion.close()
            return render_template('reiniciar_confirmar.html')
        try:
            # Obtener IDs de empleados a eliminar (lider y practicante)
            cursor.execute("SELECT id FROM empleados WHERE rol IN ('lider', 'practicante')")
            ids = [row[0] for row in cursor.fetchall()]
            if ids:
                ids_str = ','.join(str(i) for i in ids)
                # Borrar relaciones
                cursor.execute(f"DELETE FROM empleado_fortalezas WHERE empleado_id IN ({ids_str})")
                cursor.execute(f"DELETE FROM empleado_habilidades WHERE empleado_id IN ({ids_str})")
                cursor.execute(f"DELETE FROM asignaciones WHERE empleado_id IN ({ids_str})")
                # Actualizar tareas: dejar sin propietario
                cursor.execute(f"UPDATE tareas SET empleado_id = NULL WHERE empleado_id IN ({ids_str})")
                # Borrar empleados
                cursor.execute(f"DELETE FROM empleados WHERE id IN ({ids_str})")
            conexion.commit()
            flash('Base de datos reiniciada exitosamente (solo practicantes y líderes eliminados).', 'success')
        except Exception as e:
            flash(f'Error al reiniciar la base de datos: {str(e)}', 'danger')
        finally:
            conexion.close()
        return redirect(url_for('index'))
    # GET: mostrar formulario de confirmación
    return render_template('reiniciar_confirmar.html')

@app.route('/estadisticas')
@login_required
def estadisticas():
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    user_rol = current_user.rol
    user_equipo = current_user.equipo_id
    # Total de empleados (solo líderes y practicantes)
    cursor.execute("SELECT COUNT(*) as total FROM empleados WHERE rol IN ('lider', 'practicante')")
    total_empleados = cursor.fetchone()['total']
    if user_rol == 'practicante':
        # Practicante: solo ve a su líder y miembros de su equipo
        cursor.execute("""
            SELECT e.id, e.nombre, e.apellido, e.correo, e.rol, eq.nombre AS equipo
            FROM empleados e
            LEFT JOIN equipos eq ON e.equipo_id = eq.id
            WHERE e.rol = 'practicante' AND e.equipo_id = %s
            ORDER BY e.nombre, e.apellido
        """, (user_equipo,))
        lista_practicantes = cursor.fetchall()
        cursor.execute("""
            SELECT e.id, e.nombre, e.apellido, e.correo, e.rol, eq.nombre AS equipo
            FROM empleados e
            LEFT JOIN equipos eq ON e.equipo_id = eq.id
            WHERE e.rol = 'lider' AND e.equipo_id = %s
            ORDER BY e.nombre, e.apellido
        """, (user_equipo,))
        lista_lideres = cursor.fetchall()
    elif user_rol == 'lider':
        # Líder: ve a los miembros de su equipo y a sí mismo en la lista de líderes
        cursor.execute("""
            SELECT e.id, e.nombre, e.apellido, e.correo, e.rol, eq.nombre AS equipo
            FROM empleados e
            LEFT JOIN equipos eq ON e.equipo_id = eq.id
            WHERE e.rol = 'practicante' AND e.equipo_id = %s
            ORDER BY e.nombre, e.apellido
        """, (user_equipo,))
        lista_practicantes = cursor.fetchall()
        cursor.execute("""
            SELECT e.id, e.nombre, e.apellido, e.correo, e.rol, eq.nombre AS equipo
            FROM empleados e
            LEFT JOIN equipos eq ON e.equipo_id = eq.id
            WHERE e.rol = 'lider' AND e.id = %s
            ORDER BY e.nombre, e.apellido
        """, (current_user.id,))
        lista_lideres = cursor.fetchall()
    else:
        # Jefe: ve todo
        cursor.execute("""
            SELECT e.id, e.nombre, e.apellido, e.correo, e.rol, eq.nombre AS equipo
            FROM empleados e
            LEFT JOIN equipos eq ON e.equipo_id = eq.id
            WHERE e.rol = 'practicante'
            ORDER BY e.nombre, e.apellido
        """)
        lista_practicantes = cursor.fetchall()
        cursor.execute("""
            SELECT e.id, e.nombre, e.apellido, e.correo, e.rol, eq.nombre AS equipo
            FROM empleados e
            LEFT JOIN equipos eq ON e.equipo_id = eq.id
            WHERE e.rol = 'lider'
            ORDER BY e.nombre, e.apellido
        """)
        lista_lideres = cursor.fetchall()
    conexion.close()
    return render_template('estadisticas.html',
                         total_empleados=total_empleados,
                         lista_practicantes=lista_practicantes,
                         lista_lideres=lista_lideres,
                         user_rol=user_rol)

@app.route('/equipos')
def equipos():
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Obtener todos los equipos
    cursor.execute("SELECT * FROM equipos ORDER BY nombre")
    equipos = cursor.fetchall()
    # Obtener miembros por equipo (solo líderes y practicantes)
    equipos_con_miembros = []
    for equipo in equipos:
        cursor.execute("""
            SELECT * FROM empleados WHERE equipo_id = %s AND rol IN ('lider', 'practicante')
        """, (equipo['id'],))
        miembros = cursor.fetchall()
        equipos_con_miembros.append({
            'id': equipo['id'],
            'nombre': equipo['nombre'],
            'idBoard': equipo.get('idBoard'),  # Añadido el campo idBoard
            'miembros': miembros,
            'total_miembros': len(miembros)
        })
    conexion.close()
    return render_template('equipos.html', equipos=equipos_con_miembros)

def crear_tablero_trello(nombre_tablero):
    if not TRELLO_API_KEY or not TRELLO_API_TOKEN or TRELLO_API_KEY == 'TU_API_KEY' or TRELLO_API_TOKEN == 'TU_API_TOKEN':
        logging.error('API Key o Token de Trello no configurados.')
        return None, None
    url = "https://api.trello.com/1/boards/"
    params = {
        "name": nombre_tablero,
        "key": TRELLO_API_KEY,
        "token": TRELLO_API_TOKEN
    }
    response = requests.post(url, params=params)
    logging.info(f'Respuesta Trello: {response.status_code} {response.text}')
    if response.status_code == 200:
        data = response.json()
        idBoard = data['id']
        # Crear listas estándar
        listas_esenciales = ["Lista de tareas", "Pendiente", "En progreso", "Completado"]
        idListDone = None
        for nombre_lista in listas_esenciales:
            url_lista = f"https://api.trello.com/1/boards/{idBoard}/lists"
            params_lista = {"name": nombre_lista, "key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
            resp_lista = requests.post(url_lista, params=params_lista)
            if resp_lista.status_code == 200 and nombre_lista.lower() in ["completado", "hecho", "done"]:
                idListDone = resp_lista.json().get('id')
        # Obtener todas las listas del tablero
        url_lists = f"https://api.trello.com/1/boards/{idBoard}/lists"
        params_lists = {"key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
        resp_lists = requests.get(url_lists, params=params_lists)
        if resp_lists.status_code == 200:
            listas = resp_lists.json()
            # Archivar cualquier lista que no sea esencial o que sea duplicada
            nombres_esenciales = [l.lower() for l in listas_esenciales]
            listas_por_nombre = {}
            for lista in listas:
                nombre_lista = lista['name'].strip().lower()
                if nombre_lista in nombres_esenciales:
                    if nombre_lista not in listas_por_nombre:
                        listas_por_nombre[nombre_lista] = [lista['id']]
                    else:
                        listas_por_nombre[nombre_lista].append(lista['id'])
            # Archivar duplicados (dejar solo una de cada esencial)
            for nombre, ids in listas_por_nombre.items():
                for id_duplicada in ids[1:]:
                    url_archivar = f"https://api.trello.com/1/lists/{id_duplicada}/closed"
                    params_archivar = {"value": "true", "key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
                    requests.put(url_archivar, params=params_archivar)
            # Archivar cualquier lista que no sea esencial
            for lista in listas:
                nombre_lista = lista['name'].strip().lower()
                if nombre_lista not in nombres_esenciales:
                    url_archivar = f"https://api.trello.com/1/lists/{lista['id']}/closed"
                    params_archivar = {"value": "true", "key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
                    requests.put(url_archivar, params=params_archivar)
                if nombre_lista in ['completado', 'hecho', 'done']:
                    idListDone = lista['id']
            if not idListDone and listas:
                idListDone = listas[-1]['id']  # Por defecto, la última lista
        return idBoard, idListDone
    else:
        logging.error(f'Error al crear tablero: {response.text}')
        return None, None

@app.route('/equipos/nuevo', methods=['GET', 'POST'])
def agregar_equipo():
    if request.method == 'POST':
        nombre = request.form['nombre']
        try:
            idBoard, idListDone = crear_tablero_trello(nombre)
            if not idBoard:
                flash('No se pudo crear el tablero en Trello. Revisa tu API Key/Token o tu conexión a internet.', 'danger')
                return redirect(url_for('equipos'))
            conexion = get_connection()
            cursor = conexion.cursor()
            # Asegurarse de que la columna idListDone existe
            try:
                cursor.execute("ALTER TABLE equipos ADD COLUMN idListDone VARCHAR(64) NULL")
            except Exception:
                pass  # Ya existe
            cursor.execute("INSERT INTO equipos (nombre, idBoard, idListDone) VALUES (%s, %s, %s)", (nombre, idBoard, idListDone))
            conexion.commit()
            conexion.close()
            flash('Equipo y tablero de Trello creados exitosamente!', 'success')
        except Exception as e:
            logging.error(f'Error al crear equipo: {str(e)}')
            flash(f'Error al crear equipo: {str(e)}', 'danger')
        return redirect(url_for('equipos'))
    return redirect(url_for('equipos'))

@app.route('/habilidades', methods=['GET', 'POST'])
def habilidades():
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    if request.method == 'POST':
        nombre = request.form['nombre']
        descripcion = request.form.get('descripcion', '')
        try:
            cursor.execute("INSERT INTO habilidades (nombre, descripcion) VALUES (%s, %s)", (nombre, descripcion))
            conexion.commit()
            flash('Habilidad agregada exitosamente!', 'success')
        except Exception as e:
            flash(f'Error al agregar habilidad: {str(e)}', 'danger')
    cursor.execute("SELECT * FROM habilidades ORDER BY nombre")
    habilidades = cursor.fetchall()
    conexion.close()
    return render_template('habilidades.html', habilidades=habilidades)

@app.route('/competencias')
def competencias():
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Obtener solo líderes y practicantes con equipo y rol
    cursor.execute("""
        SELECT emp.id, emp.nombre, emp.apellido, emp.rol, eq.nombre AS equipo
        FROM empleados emp
        LEFT JOIN equipos eq ON emp.equipo_id = eq.id
        WHERE emp.rol IN ('lider', 'practicante')
        ORDER BY emp.nombre, emp.apellido
    """)
    empleados = cursor.fetchall()
    # Obtener habilidades y fortalezas de cada empleado
    empleados_competencias = []
    for emp in empleados:
        cursor.execute("""
            SELECT h.id, h.nombre FROM habilidades h
            JOIN empleado_habilidades eh ON h.id = eh.habilidad_id
            WHERE eh.empleado_id = %s
        """, (emp['id'],))
        habilidades = [row['nombre'] for row in cursor.fetchall()]
        cursor.execute("""
            SELECT f.id, f.nombre FROM fortalezas f
            JOIN empleado_fortalezas ef ON f.id = ef.fortaleza_id
            WHERE ef.empleado_id = %s
        """, (emp['id'],))
        fortalezas = [row['nombre'] for row in cursor.fetchall()]
        empleados_competencias.append({
            'id': emp['id'],
            'nombre': emp['nombre'],
            'apellido': emp['apellido'],
            'rol': emp['rol'],
            'equipo': emp['equipo'],
            'habilidades': habilidades,
            'fortalezas': fortalezas
        })
    # Obtener todas las habilidades y fortalezas globales
    cursor.execute("SELECT id, nombre FROM habilidades ORDER BY nombre")
    habilidades_globales = cursor.fetchall()
    cursor.execute("SELECT id, nombre FROM fortalezas ORDER BY nombre")
    fortalezas_globales = cursor.fetchall()
    conexion.close()
    return render_template('competencias.html', empleados=empleados_competencias, habilidades_globales=habilidades_globales, fortalezas_globales=fortalezas_globales)

@app.route('/competencias/editar/<int:id>', methods=['POST'])
def editar_competencias(id):
    conexion = get_connection()
    cursor = conexion.cursor()
    habilidades = request.form.getlist('habilidades')
    fortalezas = request.form.getlist('fortalezas')
    # Actualizar habilidades
    cursor.execute("DELETE FROM empleado_habilidades WHERE empleado_id=%s", (id,))
    for hab_id in habilidades:
        cursor.execute("INSERT INTO empleado_habilidades (empleado_id, habilidad_id) VALUES (%s, %s)", (id, hab_id))
    # Actualizar fortalezas
    cursor.execute("DELETE FROM empleado_fortalezas WHERE empleado_id=%s", (id,))
    for fort_id in fortalezas:
        cursor.execute("INSERT INTO empleado_fortalezas (empleado_id, fortaleza_id) VALUES (%s, %s)", (id, fort_id))
    conexion.commit()
    conexion.close()
    flash('Competencias actualizadas correctamente.', 'success')
    return redirect(url_for('competencias'))

def invitar_a_tablero_trello(email, idBoard):
    """Invita a un usuario a un tablero de Trello usando su email."""
    if not TRELLO_API_KEY or not TRELLO_API_TOKEN:
        logging.error('API Key o Token de Trello no configurados.')
        return False
    
    url = f"https://api.trello.com/1/boards/{idBoard}/members"
    params = {
        "email": email,
        "type": "normal",
        "key": TRELLO_API_KEY,
        "token": TRELLO_API_TOKEN
    }
    
    try:
        response = requests.put(url, params=params)
        if response.status_code in [200, 201]:
            logging.info(f'Usuario {email} invitado exitosamente al tablero {idBoard}')
            return True
        else:
            logging.error(f'Error al invitar usuario a Trello: {response.text}')
            return False
    except Exception as e:
        logging.error(f'Error al invitar usuario a Trello: {str(e)}')
        return False

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        correo = request.form.get('correo')
        password = request.form.get('password')
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("SELECT * FROM empleados WHERE correo = %s", (correo,))
        usuario = cursor.fetchone()
        conexion.close()
        
        if usuario and bcrypt.checkpw(password.encode('utf-8'), usuario['password'].encode('utf-8')):
            # Crear instancia de Usuario (UserMixin)
            user = Usuario()
            user.id = usuario['id']
            user.nombre = usuario['nombre']
            user.apellido = usuario['apellido']
            user.email = usuario['correo']
            user.rol = usuario['rol']
            user.telefono = usuario.get('telefono', '')
            user.password = usuario['password']
            user.equipo_id = usuario.get('equipo_id')
            
            # Autenticar con Flask-Login
            login_user(user)
            
            # Registrar sesión activa
            conexion = get_connection()
            cursor = conexion.cursor()
            cursor.execute("""
                REPLACE INTO sesiones_activas (usuario_id, rol, ultima_actividad)
                VALUES (%s, %s, %s)
            """, (user.id, user.rol, datetime.now()))
            
            # Verificar si el usuario necesita ser invitado a Trello
            if user.rol in ['lider', 'practicante'] and user.equipo_id:
                cursor.execute("""
                    SELECT et.invitado_trello, e.idBoard 
                    FROM empleados_trello et 
                    JOIN equipos e ON e.id = %s 
                    WHERE et.empleado_id = %s
                """, (user.equipo_id, user.id))
                resultado = cursor.fetchone()
                
                if resultado and not resultado[0] and resultado[1]:
                    # Invitar al usuario al tablero de Trello
                    if invitar_a_tablero_trello(user.email, resultado[1]):
                        cursor.execute("""
                            UPDATE empleados_trello 
                            SET invitado_trello = 1 
                            WHERE empleado_id = %s
                        """, (user.id,))
                        conexion.commit()
                        flash('Has sido invitado automáticamente a tu tablero de Trello.', 'success')
            
            conexion.commit()
            conexion.close()
            flash('¡Bienvenido, {}! Has iniciado sesión correctamente.'.format(usuario['nombre']), 'success')
            return redirect(url_for('index'))
        else:
            flash('Correo o contraseña incorrectos.', 'danger')
    
    # Obtener habilidades y fortalezas globales para el modal
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT * FROM habilidades ORDER BY nombre")
    habilidades_globales = cursor.fetchall()
    cursor.execute("SELECT * FROM fortalezas ORDER BY nombre")
    fortalezas_globales = cursor.fetchall()
    conexion.close()
    
    return render_template('login.html', 
                         habilidades_globales=habilidades_globales,
                         fortalezas_globales=fortalezas_globales)

@app.route('/verificar_competencias')
@login_required
def verificar_competencias():
    if current_user.rol not in ['lider', 'practicante']:
        return jsonify({'tiene_competencias': True})
    
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    
    # Verificar habilidades
    cursor.execute("""
        SELECT COUNT(*) as total FROM empleado_habilidades 
        WHERE empleado_id = %s
    """, (current_user.id,))
    tiene_habilidades = cursor.fetchone()['total'] > 0
    
    # Verificar fortalezas
    cursor.execute("""
        SELECT COUNT(*) as total FROM empleado_fortalezas 
        WHERE empleado_id = %s
    """, (current_user.id,))
    tiene_fortalezas = cursor.fetchone()['total'] > 0
    
    conexion.close()
    
    return jsonify({
        'tiene_competencias': tiene_habilidades and tiene_fortalezas
    })

@app.route('/registrar', methods=['GET', 'POST'])
def registrar():
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Obtener equipos disponibles (solo los que no tienen líder)
    cursor.execute("""
        SELECT e.* FROM equipos e
        LEFT JOIN empleados emp ON e.id = emp.equipo_id AND emp.rol = 'lider'
        WHERE emp.id IS NULL
        ORDER BY e.nombre
    """)
    equipos_disponibles = cursor.fetchall()
    # Solo permitir registro de líderes y practicantes
    roles = ['lider', 'practicante']
    if request.method == 'POST':
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        correo = request.form['correo']
        telefono = request.form.get('telefono', '')
        equipo_id = request.form.get('equipo_id')
        rol = request.form.get('rol')
        password = request.form.get('password')
        password2 = request.form.get('password2')
        if password != password2:
            flash('Las contraseñas no coinciden.', 'danger')
            return render_template('registrar.html', equipos=equipos_disponibles, roles=roles)
        if rol == 'lider' and not equipo_id:
            flash('Los líderes deben ser asignados a un equipo.', 'danger')
            return render_template('registrar.html', equipos=equipos_disponibles, roles=roles)
        # Verificar si el equipo ya tiene líder
        if rol == 'lider' and equipo_id:
            cursor.execute("""
                SELECT COUNT(*) as total FROM empleados 
                WHERE equipo_id = %s AND rol = 'lider'
            """, (equipo_id,))
            if cursor.fetchone()['total'] > 0:
                flash('Este equipo ya tiene un líder asignado.', 'danger')
                return render_template('registrar.html', equipos=equipos_disponibles, roles=roles)
        # Verificar si el correo ya existe
        cursor.execute("SELECT COUNT(*) as total FROM empleados WHERE correo = %s", (correo,))
        if cursor.fetchone()['total'] > 0:
            flash('Este correo electrónico ya está registrado.', 'danger')
            return render_template('registrar.html', equipos=equipos_disponibles, roles=roles)
        try:
            password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            cursor.execute("""
                INSERT INTO empleados 
                (nombre, apellido, correo, telefono, equipo_id, rol, password) 
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (nombre, apellido, correo, telefono, equipo_id, rol, password_hash))
            empleado_id = cursor.lastrowid
            cursor.execute("INSERT INTO empleados_trello (empleado_id, invitado_trello) VALUES (%s, 0)", (empleado_id,))
            conexion.commit()
            flash('¡Registro exitoso, {}! Ahora puedes iniciar sesión.'.format(nombre), 'success')
            return redirect(url_for('login'))
        except Exception as e:
            conexion.rollback()
            flash(f'Error al registrar: {str(e)}', 'danger')
        finally:
            conexion.close()
    return render_template('registrar.html', equipos=equipos_disponibles, roles=roles)

@app.route('/logout')
def logout():
    # Eliminar sesión activa
    try:
        conexion = get_connection()
        cursor = conexion.cursor()
        if hasattr(current_user, 'id'):
            cursor.execute("DELETE FROM sesiones_activas WHERE usuario_id = %s", (current_user.id,))
            conexion.commit()
        conexion.close()
    except Exception as e:
        pass
    logout_user()
    flash('Sesión cerrada correctamente.', 'success')
    return redirect(url_for('login'))

@app.route('/perfil')
@login_required
def perfil():
    return render_template('perfil.html', usuario=current_user)

@app.route('/actualizar_perfil', methods=['POST'])
@login_required
def actualizar_perfil():
    try:
        # Obtener datos del formulario
        nombre = request.form.get('nombre')
        apellido = request.form.get('apellido')
        email = request.form.get('email')
        telefono = request.form.get('telefono')

        # Validaciones básicas
        if not nombre or not apellido or not email:
            flash('Todos los campos obligatorios deben estar completos.', 'danger')
            return redirect(url_for('perfil'))

        # Validar formato de email
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            flash('El formato del correo electrónico no es válido.', 'danger')
            return redirect(url_for('perfil'))

        # Verificar si el email ya existe (excepto para el usuario actual)
        usuario_existente = Usuario.query.filter(
            Usuario.email == email,
            Usuario.id != current_user.id
        ).first()
        if usuario_existente:
            flash('El correo electrónico ya está registrado por otro usuario.', 'danger')
            return redirect(url_for('perfil'))

        # Actualizar datos del usuario
        current_user.nombre = nombre
        current_user.apellido = apellido
        current_user.email = email
        current_user.telefono = telefono

        db.session.commit()
        flash('Perfil actualizado exitosamente.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('Error al actualizar el perfil. Por favor, intente nuevamente.', 'danger')
        print(f"Error en actualizar_perfil: {str(e)}")

    return redirect(url_for('perfil'))

@app.route('/cambiar_password', methods=['POST'])
@login_required
def cambiar_password():
    try:
        # Obtener datos del formulario
        password_actual = request.form.get('password_actual')
        password_nueva = request.form.get('password_nueva')
        password_confirmar = request.form.get('password_confirmar')

        # Validar que la contraseña actual sea correcta
        if not check_password_hash(current_user.password, password_actual):
            flash('La contraseña actual es incorrecta.', 'danger')
            return redirect(url_for('perfil'))

        # Validar que las nuevas contraseñas coincidan
        if password_nueva != password_confirmar:
            flash('Las nuevas contraseñas no coinciden.', 'danger')
            return redirect(url_for('perfil'))

        # Validar longitud mínima de la contraseña
        if len(password_nueva) < 6:
            flash('La nueva contraseña debe tener al menos 6 caracteres.', 'danger')
            return redirect(url_for('perfil'))

        # Actualizar contraseña
        current_user.password = generate_password_hash(password_nueva)
        db.session.commit()
        flash('Contraseña actualizada exitosamente.', 'success')

    except Exception as e:
        db.session.rollback()
        flash('Error al cambiar la contraseña. Por favor, intente nuevamente.', 'danger')
        print(f"Error en cambiar_password: {str(e)}")

    return redirect(url_for('perfil'))

@app.route('/equipos/eliminar/<int:id>', methods=['POST'])
def eliminar_equipo(id):
    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)
        # Obtener equipo y correo antes de eliminar
        cursor.execute("SELECT equipo_id, correo FROM empleados WHERE id=%s", (id,))
        row = cursor.fetchone()
        equipo_id = row['equipo_id'] if row else None
        correo = row['correo'] if row else None
        # Eliminar de Trello si corresponde
        if equipo_id and correo:
            cursor.execute("SELECT idBoard FROM equipos WHERE id=%s", (equipo_id,))
            row = cursor.fetchone()
            idBoard = row['idBoard'] if row else None
            if idBoard:
                eliminar_de_tablero_trello(idBoard, correo)
        # Dejar tareas pendientes sin asignar
        cursor.execute("UPDATE tareas_trello SET empleado_id = NULL WHERE empleado_id = %s AND estado != 'completada'", (id,))
        cursor.execute("DELETE FROM empleados_trello WHERE empleado_id = %s", (id,))
        cursor.execute("DELETE FROM empleados WHERE id=%s", (id,))
        conexion.commit()
        flash('Empleado eliminado exitosamente!', 'success')
    except Exception as e:
        flash(f'Error al eliminar empleado: {str(e)}', 'danger')
    finally:
        conexion.close()
    return redirect(url_for('index'))

@app.route('/jefes', methods=['GET'])
@login_required
def jefes():
    if current_user.rol != 'jefe':
        flash('Acceso restringido solo para jefes.', 'danger')
        return redirect(url_for('index'))
    conexion = get_connection()
    cursor = conexion.cursor()
    cursor.execute("SELECT * FROM empleados WHERE rol = 'jefe' ORDER BY nombre")
    jefes = cursor.fetchall()
    # Obtener líderes y practicantes sin equipo
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT id, nombre, apellido, rol FROM empleados WHERE equipo_id IS NULL AND rol IN ('lider', 'practicante') ORDER BY nombre")
    usuarios_disponibles = cursor.fetchall()
    # Obtener equipos y si tienen líder y total de miembros
    cursor.execute("""
        SELECT eq.id, eq.nombre,
               MAX(CASE WHEN e.rol = 'lider' THEN 1 ELSE 0 END) as tiene_lider,
               SUM(CASE WHEN e.rol IN ('lider', 'practicante') THEN 1 ELSE 0 END) as total_miembros
        FROM equipos eq
        LEFT JOIN empleados e ON eq.id = e.equipo_id
        GROUP BY eq.id, eq.nombre
        ORDER BY eq.nombre
    """)
    equipos = cursor.fetchall()
    # Obtener líderes disponibles para asignar
    cursor.execute("SELECT id, nombre, apellido FROM empleados WHERE rol = 'lider' AND equipo_id IS NULL ORDER BY nombre")
    lideres_disponibles = cursor.fetchall()
    # Obtener todos los empleados líderes y practicantes con su equipo y estado de invitación
    cursor.execute("""
        SELECT e.id, e.nombre, e.apellido, e.correo, e.rol, eq.nombre as equipo, et.invitado_trello
        FROM empleados e
        LEFT JOIN equipos eq ON e.equipo_id = eq.id
        LEFT JOIN empleados_trello et ON e.id = et.empleado_id
        WHERE e.rol IN ('lider', 'practicante')
        ORDER BY FIELD(e.rol, 'jefe', 'lider', 'practicante'), e.apellido, e.nombre
    """)
    empleados_equipo = cursor.fetchall()
    # KPIs: usuarios logueados por rol
    cursor.execute("SELECT rol, COUNT(*) as total FROM sesiones_activas GROUP BY rol")
    roles_activos = {row['rol']: row['total'] for row in cursor.fetchall()}
    total_activos = sum(roles_activos.values())
    # Total empleados y equipos
    cursor.execute("SELECT COUNT(*) as total FROM empleados WHERE rol IN ('lider', 'practicante')")
    total_empleados = cursor.fetchone()['total']
    cursor.execute("SELECT COUNT(*) as total FROM equipos")
    total_equipos = cursor.fetchone()['total']
    conexion.close()
    return render_template('jefes.html', jefes=jefes, usuarios_disponibles=usuarios_disponibles, equipos=equipos, lideres_disponibles=lideres_disponibles, empleados_equipo=empleados_equipo, roles_activos=roles_activos, total_activos=total_activos, total_empleados=total_empleados, total_equipos=total_equipos)

@app.route('/jefes/crear_equipo', methods=['POST'])
@login_required
def crear_equipo_jefe():
    if current_user.rol != 'jefe':
        flash('Acceso restringido solo para jefes.', 'danger')
        return redirect(url_for('jefes'))
    nombre = request.form.get('nombre')
    miembros = request.form.getlist('miembros[]')
    if not nombre:
        flash('Debe ingresar el nombre del equipo.', 'danger')
        return redirect(url_for('jefes'))
    conexion = get_connection()
    cursor = conexion.cursor()
    try:
        cursor.execute("INSERT INTO equipos (nombre) VALUES (%s)", (nombre,))
        equipo_id = cursor.lastrowid
        for miembro_id in miembros:
            cursor.execute("UPDATE empleados SET equipo_id = %s WHERE id = %s", (equipo_id, miembro_id))
        conexion.commit()
        flash('Equipo creado correctamente.', 'success')
    except Exception as e:
        conexion.rollback()
        flash(f'Error al crear equipo: {str(e)}', 'danger')
    finally:
        conexion.close()
    return redirect(url_for('jefes'))

@app.route('/jefes/asignar_lider', methods=['POST'])
@login_required
def asignar_lider():
    if current_user.rol != 'jefe':
        flash('Acceso restringido solo para jefes.', 'danger')
        return redirect(url_for('jefes'))
    equipo_id = request.form.get('equipo_id')
    lider_id = request.form.get('lider_id')
    if not equipo_id or not lider_id:
        flash('Debe seleccionar un equipo y un líder.', 'danger')
        return redirect(url_for('jefes'))
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    try:
        # Validar que el equipo no tenga ya 2 líderes
        cursor.execute("SELECT COUNT(*) as total FROM empleados WHERE equipo_id = %s AND rol = 'lider'", (equipo_id,))
        total_lideres = cursor.fetchone()['total']
        if total_lideres >= 2:
            flash('Este equipo ya tiene 2 líderes asignados.', 'danger')
            conexion.close()
            return redirect(url_for('jefes'))
        # Obtener el usuario seleccionado
        cursor.execute("SELECT id, nombre, apellido, rol FROM empleados WHERE id = %s AND rol IN ('lider', 'practicante')", (lider_id,))
        usuario = cursor.fetchone()
        if not usuario:
            flash('Usuario no válido para ser líder.', 'danger')
            conexion.close()
            return redirect(url_for('jefes'))
        # Si es practicante, actualizar a líder
        if usuario['rol'] == 'practicante':
            cursor.execute("UPDATE empleados SET rol = 'lider', equipo_id = %s WHERE id = %s", (equipo_id, lider_id))
            conexion.commit()
            crear_y_emitir_notificacion(lider_id, 'Has sido promovido a Líder y asignado a un equipo.', 'rol')
        else:
            cursor.execute("UPDATE empleados SET equipo_id = %s WHERE id = %s AND rol = 'lider'", (equipo_id, lider_id))
            conexion.commit()
            crear_y_emitir_notificacion(lider_id, 'Has sido asignado como Líder a un equipo.', 'rol')
        flash('Líder asignado correctamente al equipo.', 'success')
    except Exception as e:
        conexion.rollback()
        flash(f'Error al asignar líder: {str(e)}', 'danger')
    finally:
        conexion.close()
    return redirect(url_for('jefes'))

@app.route('/equipos/integrantes/<int:equipo_id>')
def integrantes_equipo(equipo_id):
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("""
        SELECT nombre, apellido, correo, rol FROM empleados WHERE equipo_id = %s AND rol IN ('lider', 'practicante') ORDER BY rol, nombre
    """, (equipo_id,))
    integrantes = cursor.fetchall()
    conexion.close()
    return jsonify({'integrantes': integrantes})

@app.route('/jefes/editar_equipo', methods=['POST'])
@login_required
def editar_equipo():
    if current_user.rol != 'jefe':
        flash('Acceso restringido solo para jefes.', 'danger')
        return redirect(url_for('jefes'))
    equipo_id = request.form.get('equipo_id')
    nombre = request.form.get('nombre')
    lider_id = request.form.get('lider_id')
    miembros = request.form.getlist('miembros[]')
    if not equipo_id or not nombre:
        flash('Faltan datos obligatorios.', 'danger')
        return redirect(url_for('jefes'))
    conexion = get_connection()
    cursor = conexion.cursor()
    try:
        # Actualizar nombre del equipo
        cursor.execute("UPDATE equipos SET nombre = %s WHERE id = %s", (nombre, equipo_id))
        # Asignar/cambiar líder
        if lider_id:
            # Validar que el líder no esté en otro equipo
            cursor.execute("SELECT equipo_id FROM empleados WHERE id = %s AND rol = 'lider'", (lider_id,))
            lider_row = cursor.fetchone()
            if lider_row and lider_row[0] and int(lider_row[0]) != int(equipo_id):
                flash('El líder seleccionado ya pertenece a otro equipo.', 'danger')
                conexion.close()
                return redirect(url_for('jefes'))
            # Asignar líder a este equipo
            cursor.execute("UPDATE empleados SET equipo_id = %s WHERE id = %s AND rol = 'lider'", (equipo_id, lider_id))
        # Actualizar miembros (líderes/practicantes)
        # Primero, quitar a todos los miembros actuales de este equipo
        cursor.execute("UPDATE empleados SET equipo_id = NULL WHERE equipo_id = %s AND rol IN ('lider', 'practicante')", (equipo_id,))
        # Luego, asignar los seleccionados
        for miembro_id in miembros:
            cursor.execute("UPDATE empleados SET equipo_id = %s WHERE id = %s AND rol IN ('lider', 'practicante')", (equipo_id, miembro_id))
        conexion.commit()
        flash('Equipo actualizado correctamente.', 'success')
    except Exception as e:
        conexion.rollback()
        flash(f'Error al editar equipo: {str(e)}', 'danger')
    finally:
        conexion.close()
    return redirect(url_for('jefes'))

def limpiar_sesiones_inactivas():
    conexion = get_connection()
    cursor = conexion.cursor()
    cursor.execute("""
        DELETE FROM sesiones_activas
        WHERE ultima_actividad < (NOW() - INTERVAL 30 MINUTE)
    """)
    conexion.commit()
    conexion.close()

@app.route('/exportar_empleados_excel')
@login_required
@rol_requerido('jefe')
def exportar_empleados_excel():
    conexion = get_connection()
    cursor = conexion.cursor()
    # Obtener todos los empleados relevantes
    cursor.execute("""
        SELECT e.id, e.nombre, e.apellido, e.correo, e.rol, eq.nombre as equipo, e.telefono, e.horas_disponibles, e.habilidades
        FROM empleados e
        LEFT JOIN equipos eq ON e.equipo_id = eq.id
        WHERE e.rol IN ('lider', 'practicante')
        ORDER BY FIELD(e.rol, 'jefe', 'lider', 'practicante'), e.apellido, e.nombre
    """)
    empleados = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados"
    ws.append([
        "Nombre", "Apellido", "Correo", "Rol", "Equipo", "Teléfono", "Habilidades", "Fortalezas", "Horas Disponibles",
        "Tareas Asignadas", "Tareas Completadas", "Tareas Pendientes", "Prioridad Más Alta", "Títulos de Tareas"
    ])

    for emp in empleados:
        empleado_id = emp[0]
        habilidades, fortalezas = obtener_habilidades_fortalezas(cursor, empleado_id)
        # Tareas asignadas
        cursor.execute("SELECT COUNT(*) FROM tareas WHERE empleado_id = %s", (empleado_id,))
        total_tareas = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM tareas WHERE empleado_id = %s AND estado = 'completada'", (empleado_id,))
        tareas_completadas = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM tareas WHERE empleado_id = %s AND estado != 'completada'", (empleado_id,))
        tareas_pendientes = cursor.fetchone()[0]
        # Prioridad más alta
        cursor.execute("SELECT prioridad FROM tareas WHERE empleado_id = %s", (empleado_id,))
        prioridades = [row[0] for row in cursor.fetchall()]
        prioridad_mas_alta = 'N/A'
        if 'alta' in prioridades:
            prioridad_mas_alta = 'alta'
        elif 'media' in prioridades:
            prioridad_mas_alta = 'media'
        elif 'baja' in prioridades:
            prioridad_mas_alta = 'baja'
        # Títulos de tareas
        cursor.execute("SELECT titulo FROM tareas WHERE empleado_id = %s", (empleado_id,))
        titulos_tareas = ', '.join([row[0] for row in cursor.fetchall()])
        ws.append([
            emp[1], emp[2], emp[3], emp[4], emp[5], emp[6],
            habilidades or emp[8] or '', fortalezas, emp[7],
            total_tareas, tareas_completadas, tareas_pendientes, prioridad_mas_alta, titulos_tareas
        ])
    # Ajustar el ancho de las columnas automáticamente según el contenido
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Un poco de espacio extra
        ws.column_dimensions[column].width = adjusted_width
    cursor.close()
    conexion.close()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="empleados_completo.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/exportar_empleados_pdf')
@login_required
@rol_requerido('jefe')
def exportar_empleados_pdf():
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.lib.utils import ImageReader
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=rl_landscape(A4))
    width, height = rl_landscape(A4)
    y = height - 60
    # Logo centrado
    logo_path = os.path.join('proyecto_empleados', 'static', 'img', 'Reflexo_Perulogo.png')
    if os.path.exists(logo_path):
        logo_width = 120
        logo_height = 80
        c.drawImage(ImageReader(logo_path), (width - logo_width) / 2, y, width=logo_width, height=logo_height, mask='auto')
        y -= logo_height + 10
    # Título principal centrado
    c.setFont('Helvetica-Bold', 22)
    c.setFillColor(colors.HexColor('#4F81BD'))
    c.drawCentredString(width / 2, y, "APM INVERSIONES EIRL")
    c.setFont('Helvetica', 13)
    c.setFillColor(colors.black)
    y -= 30
    c.drawCentredString(width / 2, y, "Socio estratégico en el crecimiento de clínicas y profesionales de la salud")
    y -= 40
    # Definir columnas
    col1_x = width * 0.08
    col2_x = width * 0.54
    col_width = width * 0.38
    y_start = y
    y1 = y_start
    y2 = y_start
    # Función para bloques en columna
    def draw_block_col(x, y, title, text, color, font_size=15, text_font='Helvetica', text_size=11, space=28):
        c.setFont('Helvetica-Bold', font_size)
        c.setFillColor(color)
        c.drawString(x, y, title)
        c.setFont(text_font, text_size)
        c.setFillColor(colors.black)
        y -= 18
        lines = split_text(text, col_width-10, c, text_font, text_size)
        for line in lines:
            c.drawString(x, y, line)
            y -= 15
        return y - space
    def split_text(text, max_width, canvas, font, size):
        words = text.split()
        lines = []
        line = ''
        canvas.setFont(font, size)
        for word in words:
            test_line = f'{line} {word}'.strip()
            if canvas.stringWidth(test_line) <= max_width:
                line = test_line
            else:
                lines.append(line)
                line = word
        if line:
            lines.append(line)
        return lines
    # Columna 1
    y1 = draw_block_col(col1_x, y1, "Misión:", "Ser el socio estratégico en el crecimiento de clínicas y profesionales de la salud, implementando soluciones de marketing digital personalizadas e integración de tecnologías CRM y ERP.", colors.HexColor('#00CFFF'))
    y1 = draw_block_col(col1_x, y1, "Objetivo:", "Lograr que clínicas y profesionales experimenten un crecimiento constante y sostenible mediante estrategias digitales y tecnológicas avanzadas.", colors.HexColor('#FFB347'))
    y1 = draw_block_col(col1_x, y1, "Valores:", "Honestidad, Innovación, Respeto, Calidad", colors.HexColor('#7F9CF5'))
    y1 = draw_block_col(col1_x, y1, "Servicios principales:", "- Desarrollo y mantenimiento web\n- Marketing digital (Google Ads, Facebook Ads, SEO)\n- Integración CRM y ERP\n- Publicidad en redes sociales", colors.HexColor('#00CFFF'))
    # Columna 2
    y2 = draw_block_col(col2_x, y2, "Visión:", "Liderar la transformación digital del sector salud, impulsando el crecimiento sostenible de clínicas y profesionales visionarios con soluciones tecnológicas y de marketing innovadoras.", colors.HexColor('#7F9CF5'))
    y2 = draw_block_col(col2_x, y2, "¿A quién va dirigido?", "- Clínicas consolidadas y en expansión\n- Profesionales de la salud independientes\n- Redes de clínicas/franquicias\n- Instituciones de salud", colors.HexColor('#00CFFF'))
    y2 = draw_block_col(col2_x, y2, "Sobre el sistema:", "Esta plataforma digitaliza y optimiza la gestión de empleados, equipos y tareas en clínicas y empresas de salud, integrando herramientas modernas como Trello y paneles de control avanzados para impulsar la eficiencia y el crecimiento.", colors.HexColor('#7F9CF5'), font_size=15, text_font='Helvetica', text_size=11, space=18)
    # Frases motivacionales centradas abajo
    y_frases = min(y1, y2) - 30
    frases = [
        ("La tecnología en salud no solo transforma clínicas, transforma vidas.", colors.HexColor('#00CFFF')),
        ("Innovar es cuidar mejor.", colors.HexColor('#43e97b')),
        ("El futuro de la salud es digital, seguro y humano.", colors.HexColor('#00CFFF'))
    ]
    c.setFont('Helvetica-Oblique', 12)
    for frase, color in frases:
        c.setFillColor(color)
        c.drawCentredString(width / 2, y_frases, f'"{frase}"')
        y_frases -= 18
    c.save()
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="empresa_presentacion.pdf",
        mimetype="application/pdf"
    )

@app.route('/reportes', methods=['GET', 'POST'])
@login_required
def reportes():
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Obtener filtros
    equipos = []
    empleados = []
    cursor.execute("SELECT id, nombre FROM equipos ORDER BY nombre")
    equipos = cursor.fetchall()
    cursor.execute("SELECT id, nombre, apellido FROM empleados ORDER BY nombre")
    empleados = cursor.fetchall()
    # Filtros del usuario
    equipo_id = request.args.get('equipo_id')
    empleado_id = request.args.get('empleado_id')
    estado = request.args.get('estado')
    prioridad = request.args.get('prioridad')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    # Construir consulta dinámica
    sql = "SELECT t.*, e.nombre as empleado_nombre, e.apellido as empleado_apellido, eq.nombre as equipo_nombre FROM tareas_trello t LEFT JOIN empleados e ON t.empleado_id = e.id LEFT JOIN equipos eq ON e.equipo_id = eq.id WHERE 1=1"
    params = []
    if equipo_id:
        sql += " AND eq.id = %s"
        params.append(equipo_id)
    if empleado_id:
        sql += " AND e.id = %s"
        params.append(empleado_id)
    if estado:
        sql += " AND t.estado = %s"
        params.append(estado)
    if prioridad:
        sql += " AND t.prioridad = %s"
        params.append(prioridad)
    if fecha_inicio:
        sql += " AND t.fecha_creacion >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        sql += " AND t.fecha_creacion <= %s"
        params.append(fecha_fin)
    sql += " ORDER BY t.fecha_creacion DESC"
    cursor.execute(sql, tuple(params))
    tareas = cursor.fetchall()
    conexion.close()
    return render_template('reportes.html', equipos=equipos, empleados=empleados, tareas=tareas, filtros={
        'equipo_id': equipo_id, 'empleado_id': empleado_id, 'estado': estado, 'prioridad': prioridad, 'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin
    })

@app.route('/reportes/exportar_excel')
@login_required
def exportar_reporte_excel():
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Obtener filtros igual que en /reportes
    equipo_id = request.args.get('equipo_id')
    empleado_id = request.args.get('empleado_id')
    estado = request.args.get('estado')
    prioridad = request.args.get('prioridad')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    sql = "SELECT t.*, e.nombre as empleado_nombre, e.apellido as empleado_apellido, eq.nombre as equipo_nombre FROM tareas_trello t LEFT JOIN empleados e ON t.empleado_id = e.id LEFT JOIN equipos eq ON e.equipo_id = eq.id WHERE 1=1"
    params = []
    if equipo_id:
        sql += " AND eq.id = %s"
        params.append(equipo_id)
    if empleado_id:
        sql += " AND e.id = %s"
        params.append(empleado_id)
    if estado:
        sql += " AND t.estado = %s"
        params.append(estado)
    if prioridad:
        sql += " AND t.prioridad = %s"
        params.append(prioridad)
    if fecha_inicio:
        sql += " AND t.fecha_creacion >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        sql += " AND t.fecha_creacion <= %s"
        params.append(fecha_fin)
    sql += " ORDER BY t.fecha_creacion DESC"
    cursor.execute(sql, tuple(params))
    tareas = cursor.fetchall()
    conexion.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Tareas"
    ws.append(["Título", "Descripción", "Empleado", "Equipo", "Estado", "Prioridad", "Fecha Creación", "Fecha Límite", "Tiempo Estimado", "Tiempo Real"])
    for t in tareas:
        ws.append([
            t['titulo'], t['descripcion'], f"{t['empleado_nombre']} {t['empleado_apellido']}", t['equipo_nombre'],
            t['estado'], t['prioridad'],
            t['fecha_creacion'].strftime('%d/%m/%Y') if t['fecha_creacion'] else '',
            t['fecha_limite'].strftime('%d/%m/%Y') if t['fecha_limite'] else '',
            t['tiempo_estimado'], t['tiempo_real']
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="reporte_tareas.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/reportes/exportar_pdf')
@login_required
def exportar_reporte_pdf():
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Obtener filtros igual que en /reportes
    equipo_id = request.args.get('equipo_id')
    empleado_id = request.args.get('empleado_id')
    estado = request.args.get('estado')
    prioridad = request.args.get('prioridad')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    sql = "SELECT t.*, e.nombre as empleado_nombre, e.apellido as empleado_apellido, eq.nombre as equipo_nombre FROM tareas_trello t LEFT JOIN empleados e ON t.empleado_id = e.id LEFT JOIN equipos eq ON e.equipo_id = eq.id WHERE 1=1"
    params = []
    if equipo_id:
        sql += " AND eq.id = %s"
        params.append(equipo_id)
    if empleado_id:
        sql += " AND e.id = %s"
        params.append(empleado_id)
    if estado:
        sql += " AND t.estado = %s"
        params.append(estado)
    if prioridad:
        sql += " AND t.prioridad = %s"
        params.append(prioridad)
    if fecha_inicio:
        sql += " AND t.fecha_creacion >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        sql += " AND t.fecha_creacion <= %s"
        params.append(fecha_fin)
    sql += " ORDER BY t.fecha_creacion DESC"
    cursor.execute(sql, tuple(params))
    tareas = cursor.fetchall()
    conexion.close()
    data = [["Título", "Descripción", "Empleado", "Equipo", "Estado", "Prioridad", "Fecha Creación", "Fecha Límite", "Tiempo Estimado", "Tiempo Real"]]
    for t in tareas:
        data.append([
            t['titulo'], t['descripcion'], f"{t['empleado_nombre']} {t['empleado_apellido']}", t['equipo_nombre'],
            t['estado'], t['prioridad'],
            t['fecha_creacion'].strftime('%d/%m/%Y') if t['fecha_creacion'] else '',
            t['fecha_limite'].strftime('%d/%m/%Y') if t['fecha_limite'] else '',
            t['tiempo_estimado'], t['tiempo_real']
        ])
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
    ]))
    table_width, table_height = table.wrapOn(c, width, height)
    table.drawOn(c, 20, height - table_height - 40)
    c.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="reporte_tareas.pdf", mimetype="application/pdf")

@app.route('/equipos/<int:equipo_id>')
@login_required
@equipo_requerido
def vista_equipo(equipo_id):
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Obtener datos del equipo
    cursor.execute("SELECT * FROM equipos WHERE id = %s", (equipo_id,))
    equipo = cursor.fetchone()
    if not equipo:
        conexion.close()
        flash('Equipo no encontrado.', 'danger')
        return redirect(url_for('equipos'))
    # Obtener miembros del equipo
    cursor.execute("SELECT * FROM empleados WHERE equipo_id = %s AND rol IN ('lider', 'practicante') ORDER BY rol, nombre", (equipo_id,))
    miembros = cursor.fetchall()
    # Obtener tareas del equipo
    cursor.execute("""
        SELECT t.*, e.nombre as asignado_nombre, e.apellido as asignado_apellido
        FROM tareas_trello t
        LEFT JOIN empleados e ON t.empleado_id = e.id
        WHERE t.equipo_id = %s
        ORDER BY t.fecha_limite ASC
    """, (equipo_id,))
    tareas = cursor.fetchall()
    # Obtener mensajes del equipo (simulado, puedes adaptar a tu modelo real)
    mensajes = []
    conexion.close()
    return render_template('equipos/vista_equipo.html', equipo=equipo, miembros=miembros, tareas=tareas, mensajes=mensajes)

# Evento de conexión
@socketio.on('connect')
def handle_connect():
    print('Usuario conectado')
    equipo_id = None
    try:
        # Si el usuario está autenticado y tiene equipo
        if current_user.is_authenticated and hasattr(current_user, 'equipo_id') and current_user.equipo_id:
            equipo_id = str(current_user.equipo_id)
            from flask_socketio import join_room
            join_room(f'equipo_{equipo_id}')
    except Exception as e:
        print('Error al unir a sala de equipo:', e)
    emit('usuario_conectado', {'hora': datetime.now().strftime('%H:%M')}, broadcast=True)

# Evento de desconexión
@socketio.on('disconnect')
def handle_disconnect():
    print('Usuario desconectado')
    emit('usuario_desconectado', {'hora': datetime.now().strftime('%H:%M')}, broadcast=True)

# Evento de mensaje global
@socketio.on('mensaje')
def handle_mensaje(data):
    mensaje = data.get('mensaje', '').strip()
    usuario = data.get('usuario', 'Invitado')
    if not mensaje or len(mensaje) > 200:
        return
    hora = datetime.now().strftime('%H:%M')
    emit('mensaje', {'usuario': usuario, 'mensaje': mensaje, 'hora': hora}, broadcast=True)

# Evento de mensaje de equipo
@socketio.on('mensaje_equipo')
def handle_mensaje_equipo(data):
    mensaje = data.get('mensaje', '').strip()
    usuario = data.get('usuario', 'Invitado')
    equipo_id = data.get('equipo_id')
    if not mensaje or len(mensaje) > 200 or not equipo_id:
        return
    hora = datetime.now().strftime('%H:%M')
    emit('mensaje_equipo', {'usuario': usuario, 'mensaje': mensaje, 'hora': hora}, room=f'equipo_{equipo_id}')

@app.route('/trello')
def trello_panel():
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Obtener todos los equipos con tablero Trello
    cursor.execute("SELECT * FROM equipos WHERE idBoard IS NOT NULL ORDER BY nombre")
    equipos = cursor.fetchall()
    equipos_con_tareas = []
    for equipo in equipos:
        # Obtener miembros del equipo
        cursor.execute("SELECT id, nombre, apellido FROM empleados WHERE equipo_id = %s AND rol IN ('lider', 'practicante') ORDER BY nombre", (equipo['id'],))
        miembros = cursor.fetchall()
        # Obtener tareas del equipo
        cursor.execute("""
            SELECT t.*, e.nombre as asignado_nombre, e.apellido as asignado_apellido
            FROM tareas_trello t
            LEFT JOIN empleados e ON t.empleado_id = e.id
            WHERE t.equipo_id = %s
            ORDER BY t.fecha_limite ASC
        """, (equipo['id'],))
        tareas = cursor.fetchall()
        equipos_con_tareas.append({
            'id': equipo['id'],
            'nombre': equipo['nombre'],
            'idBoard': equipo['idBoard'],
            'miembros': miembros,
            'tareas': tareas
        })
    conexion.close()
    return render_template('trello/panel.html', equipos=equipos_con_tareas)

@app.route('/trello/miembros/<int:equipo_id>', methods=['GET', 'POST'])
def gestionar_miembros_equipo(equipo_id):
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Obtener datos del equipo
    cursor.execute("SELECT * FROM equipos WHERE id = %s", (equipo_id,))
    equipo = cursor.fetchone()
    # Obtener miembros actuales
    cursor.execute("SELECT id, nombre, apellido, rol FROM empleados WHERE equipo_id = %s AND rol IN ('lider', 'practicante') ORDER BY nombre", (equipo_id,))
    miembros = cursor.fetchall()
    # Obtener usuarios disponibles (no asignados a ningún equipo)
    cursor.execute("SELECT id, nombre, apellido, rol FROM empleados WHERE equipo_id IS NULL AND rol IN ('lider', 'practicante') ORDER BY nombre", ())
    disponibles = cursor.fetchall()
    if request.method == 'POST':
        quitar_id = request.form.get('quitar_miembro_id')
        if quitar_id:
            cursor2 = conexion.cursor()
            cursor2.execute("UPDATE empleados SET equipo_id = NULL WHERE id = %s", (quitar_id,))
            conexion.commit()
            cursor2.close()
            conexion.close()
            flash('Miembro quitado correctamente.', 'success')
            return redirect(url_for('gestionar_miembros_equipo', equipo_id=equipo_id))
        nuevos_miembros = request.form.getlist('miembros')
        # Quitar todos los miembros actuales
        cursor2 = conexion.cursor()
        cursor2.execute("UPDATE empleados SET equipo_id = NULL WHERE equipo_id = %s AND rol IN ('lider', 'practicante')", (equipo_id,))
        # Asignar los seleccionados
        for miembro_id in nuevos_miembros:
            cursor2.execute("UPDATE empleados SET equipo_id = %s WHERE id = %s", (equipo_id, miembro_id))
        conexion.commit()
        cursor2.close()
        conexion.close()
        flash('Miembros actualizados correctamente.', 'success')
        return redirect(url_for('gestionar_miembros_equipo', equipo_id=equipo_id))
    conexion.close()
    return render_template('trello/gestionar_miembros.html', equipo=equipo, miembros=miembros, disponibles=disponibles)

@app.route('/equipos/sincronizar_trello', methods=['POST'])
@login_required
def sincronizar_equipos_trello():
    if current_user.rol != 'jefe':
        flash('Acceso restringido solo para jefes.', 'danger')
        return redirect(url_for('equipos'))
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # 1. Importar tableros existentes en Trello
    url = f"https://api.trello.com/1/members/me/boards"
    params = {"key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN, "fields": "name"}
    response = requests.get(url, params=params)
    tableros = response.json() if response.status_code == 200 else []
    # Obtener todos los idBoard ya registrados
    cursor.execute("SELECT idBoard FROM equipos WHERE idBoard IS NOT NULL AND idBoard != ''")
    idboards_local = set(row['idBoard'] for row in cursor.fetchall())
    total_importados = 0
    for tablero in tableros:
        idBoard = tablero.get('id')
        nombre = tablero.get('name')
        if idBoard:
            # Verificar si el idBoard ya existe
            if idBoard in idboards_local:
                continue
            # Verificar si el nombre ya existe y agregar sufijo si es necesario
            nombre_final = nombre
            sufijo = 2
            while True:
                cursor.execute("SELECT COUNT(*) as total FROM equipos WHERE nombre = %s", (nombre_final,))
                if cursor.fetchone()['total'] == 0:
                    break
                nombre_final = f"{nombre} ({sufijo})"
                sufijo += 1
            cursor.execute("INSERT INTO equipos (nombre, idBoard) VALUES (%s, %s)", (nombre_final, idBoard))
            conexion.commit()
            total_importados += 1
    # 2. Lógica original: crear tableros en Trello para equipos locales sin tablero
    cursor.execute("SELECT * FROM equipos WHERE idBoard IS NULL OR idBoard = ''")
    equipos_sin_trello = cursor.fetchall()
    total_creados = 0
    for equipo in equipos_sin_trello:
        idBoard, idListDone = crear_tablero_trello(equipo['nombre'])
        if idBoard:
            cursor2 = conexion.cursor()
            cursor2.execute("UPDATE equipos SET idBoard = %s, idListDone = %s WHERE id = %s", (idBoard, idListDone, equipo['id']))
            conexion.commit()
            cursor2.close()
            total_creados += 1
    conexion.close()
    flash(f'Se importaron {total_importados} tableros de Trello y se sincronizaron {total_creados} equipos locales.', 'success')
    return redirect(url_for('equipos'))

@app.route('/equipos/editar_trello/<int:equipo_id>', methods=['GET', 'POST'])
@login_required
def editar_equipo_trello(equipo_id):
    if current_user.rol != 'jefe':
        flash('Acceso restringido solo para jefes.', 'danger')
        return redirect(url_for('equipos'))
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT * FROM equipos WHERE id = %s", (equipo_id,))
    equipo = cursor.fetchone()
    if not equipo:
        conexion.close()
        flash('Equipo no encontrado.', 'danger')
        return redirect(url_for('equipos'))
    if request.method == 'POST':
        nuevo_nombre = request.form.get('nombre')
        if nuevo_nombre and nuevo_nombre != equipo['nombre']:
            # Actualizar en Trello
            if equipo['idBoard']:
                url = f"https://api.trello.com/1/boards/{equipo['idBoard']}"
                params = {"name": nuevo_nombre, "key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
                resp = requests.put(url, params=params)
                if resp.status_code != 200:
                    flash('No se pudo actualizar el nombre en Trello.', 'danger')
            # Actualizar en BD
            cursor.execute("UPDATE equipos SET nombre = %s WHERE id = %s", (nuevo_nombre, equipo_id))
            conexion.commit()
            flash('Nombre de equipo actualizado correctamente.', 'success')
        conexion.close()
        return redirect(url_for('equipos'))
    conexion.close()
    return render_template('equipos/editar_trello.html', equipo=equipo)

@app.route('/equipos/eliminar_trello/<int:equipo_id>', methods=['POST'])
@login_required
def eliminar_equipo_trello(equipo_id):
    if current_user.rol != 'jefe':
        return jsonify(success=False, message='Acceso restringido solo para jefes.'), 403
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT * FROM equipos WHERE id = %s", (equipo_id,))
    equipo = cursor.fetchone()
    if not equipo:
        conexion.close()
        return jsonify(success=False, message='Equipo no encontrado.'), 404
    # Eliminar en Trello
    if equipo['idBoard']:
        url = f"https://api.trello.com/1/boards/{equipo['idBoard']}"
        params = {"key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
        resp = requests.delete(url, params=params)
        if resp.status_code not in [200, 202, 204]:
            conexion.close()
            return jsonify(success=False, message='No se pudo eliminar el tablero en Trello.'), 500
    # Eliminar en BD
    try:
        cursor.execute("DELETE FROM equipos WHERE id = %s", (equipo_id,))
        conexion.commit()
        conexion.close()
        return jsonify(success=True)
    except Exception as e:
        conexion.rollback()
        conexion.close()
        return jsonify(success=False, message=f'Error al eliminar en la base de datos: {str(e)}'), 500

@app.route('/empleados/resetear_invitacion_trello/<int:empleado_id>', methods=['POST'])
@login_required
@rol_requerido('jefe')
def resetear_invitacion_trello(empleado_id):
    conexion = get_connection()
    cursor = conexion.cursor()
    try:
        cursor.execute("UPDATE empleados_trello SET invitado_trello = 0 WHERE empleado_id = %s", (empleado_id,))
        conexion.commit()
        flash('La invitación a Trello ha sido reseteada. El usuario será invitado nuevamente al ingresar.', 'success')
    except Exception as e:
        conexion.rollback()
        flash(f'Error al resetear invitación: {str(e)}', 'danger')
    finally:
        conexion.close()
    return redirect(url_for('index'))

def inicializar_tabla_empleados_trello():
    """Inicializa la tabla empleados_trello si no existe."""
    conexion = get_connection()
    cursor = conexion.cursor()
    try:
        # Crear la tabla si no existe
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS empleados_trello (
                empleado_id INT PRIMARY KEY,
                invitado_trello BOOLEAN DEFAULT 0,
                FOREIGN KEY (empleado_id) REFERENCES empleados(id) ON DELETE CASCADE
            )
        """)
        
        # Insertar registros para empleados que no tengan uno
        cursor.execute("""
            INSERT IGNORE INTO empleados_trello (empleado_id, invitado_trello)
            SELECT id, 0 FROM empleados
            WHERE id NOT IN (SELECT empleado_id FROM empleados_trello)
        """)
        
        conexion.commit()
        logging.info('Tabla empleados_trello inicializada correctamente')
    except Exception as e:
        logging.error(f'Error al inicializar tabla empleados_trello: {str(e)}')
        conexion.rollback()
    finally:
        cursor.close()
        conexion.close()

# Inicializar la tabla al arrancar la aplicación
inicializar_tabla_empleados_trello()

# Función auxiliar para obtener el idMember de Trello por correo
def obtener_id_member_trello(email):
    url = f"https://api.trello.com/1/members/{email}"
    params = {"key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
    resp = requests.get(url, params=params)
    if resp.status_code == 200:
        return resp.json().get('id')
    return None

# Función para eliminar un miembro de un tablero Trello
def eliminar_de_tablero_trello(idBoard, email):
    idMember = obtener_id_member_trello(email)
    if not idMember:
        return False
    url = f"https://api.trello.com/1/boards/{idBoard}/members/{idMember}"
    params = {"key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
    resp = requests.delete(url, params=params)
    return resp.status_code in [200, 204]

@app.route('/jefes/crear_jefe', methods=['POST'])
@login_required
def crear_jefe():
    nombre = request.form.get('nombre')
    apellido = request.form.get('apellido')
    correo = request.form.get('correo')
    password = request.form.get('password')

    if not (nombre and apellido and correo and password):
        flash('Todos los campos son obligatorios.', 'danger')
        return redirect(url_for('jefes'))

    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Validar límite de jefes (máximo 4 por ahora)
    cursor.execute("SELECT COUNT(*) as total FROM empleados WHERE rol = 'jefe'")
    total_jefes = cursor.fetchone()['total']
    if total_jefes >= 4:
        flash('No se pueden registrar más de 4 jefes. Elimine uno para agregar otro.', 'danger')
        conexion.close()
        return redirect(url_for('jefes'))
    # Validar correo único
    cursor.execute("SELECT COUNT(*) as total FROM empleados WHERE correo = %s", (correo,))
    if cursor.fetchone()['total'] > 0:
        flash('El correo electrónico ya está registrado.', 'danger')
        conexion.close()
        return redirect(url_for('jefes'))
    try:
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        cursor.execute("""
            INSERT INTO empleados (nombre, apellido, correo, rol, password)
            VALUES (%s, %s, %s, 'jefe', %s)
        """, (nombre, apellido, correo, password_hash))
        conexion.commit()
        flash('Jefe creado exitosamente.', 'success')
    except Exception as e:
        conexion.rollback()
        flash(f'Error al crear jefe: {str(e)}', 'danger')
    finally:
        conexion.close()
    return redirect(url_for('jefes'))

@app.route('/equipos/normalizar_listas_trello', methods=['POST'])
@login_required
def normalizar_listas_trello():
    if current_user.rol != 'jefe':
        flash('Acceso restringido solo para jefes.', 'danger')
        return redirect(url_for('equipos'))
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT id, nombre, idBoard FROM equipos WHERE idBoard IS NOT NULL")
    equipos = cursor.fetchall()
    total_normalizados = 0
    listas_esenciales = ["lista de tareas", "pendiente", "en progreso", "completado"]
    for equipo in equipos:
        idBoard = equipo['idBoard']
        # Obtener listas actuales
        url_lists = f"https://api.trello.com/1/boards/{idBoard}/lists"
        params = {"key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
        resp = requests.get(url_lists, params=params)
        if resp.status_code != 200:
            continue
        listas_actuales = resp.json()
        nombres_actuales = [l['name'].strip().lower() for l in listas_actuales]
        # Eliminar listas que no sean esenciales
        for lista in listas_actuales:
            nombre_lista = lista['name'].strip().lower()
            if nombre_lista not in listas_esenciales:
                # Archivar la lista (Trello no permite eliminar, solo archivar)
                url_archivar = f"https://api.trello.com/1/lists/{lista['id']}/closed"
                params_archivar = {"value": "true", "key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
                requests.put(url_archivar, params=params_archivar)
        # Crear las listas esenciales que falten
        for nombre_lista in listas_esenciales:
            if nombre_lista not in nombres_actuales:
                url_lista = f"https://api.trello.com/1/boards/{idBoard}/lists"
                params_lista = {"name": nombre_lista, "key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
                resp_lista = requests.post(url_lista, params=params_lista)
                if resp_lista.status_code == 200:
                    total_normalizados += 1
    conexion.close()
    flash(f'Se normalizaron las listas en {total_normalizados} tableros/equipos.', 'success')
    return redirect(url_for('equipos'))

@app.route('/api/notificaciones', methods=['GET'])
@login_required
def get_notificaciones():
    notifs = Notification.query.filter_by(usuario_id=current_user.id).order_by(Notification.fecha.desc()).limit(20).all()
    return jsonify([
        {
            'id': n.id,
            'mensaje': n.mensaje,
            'tipo': n.tipo,
            'leido': n.leido,
            'fecha': n.fecha.strftime('%Y-%m-%d %H:%M')
        } for n in notifs
    ])

@app.route('/api/notificaciones/<int:notif_id>/leer', methods=['POST'])
@login_required
def marcar_notificacion_leida(notif_id):
    notif = Notification.query.filter_by(id=notif_id, usuario_id=current_user.id).first_or_404()
    notif.leido = True
    from . import db
    db.session.commit()
    # Emitir evento por socket al usuario
    socketio.emit(f'notificacion_leida_{current_user.id}', {'id': notif.id})
    return jsonify({'success': True})

@app.route('/api/notificaciones/<int:notif_id>', methods=['DELETE'])
@login_required
def eliminar_notificacion(notif_id):
    notif = Notification.query.filter_by(id=notif_id, usuario_id=current_user.id).first_or_404()
    from . import db
    db.session.delete(notif)
    db.session.commit()
    # Emitir evento por socket al usuario
    socketio.emit(f'notificacion_eliminada_{current_user.id}', {'id': notif.id})
    return jsonify({'success': True})

@app.route('/subir_avatar', methods=['POST'])
@login_required
def subir_avatar():
    file = request.files.get('avatar')
    if not file:
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('perfil'))
    if file.filename == '':
        flash('Nombre de archivo inválido.', 'danger')
        return redirect(url_for('perfil'))
    if not (file.filename.lower().endswith('.jpg') or file.filename.lower().endswith('.jpeg') or file.filename.lower().endswith('.png')):
        flash('Solo se permiten imágenes JPG o PNG.', 'danger')
        return redirect(url_for('perfil'))
    if file.content_length and file.content_length > 1*1024*1024:
        flash('La imagen no debe superar 1MB.', 'danger')
        return redirect(url_for('perfil'))
    try:
        from PIL import Image, UnidentifiedImageError
        filename = f"avatars/{current_user.id}.png"
        path = os.path.join('proyecto_empleados/static/img', filename)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        # Procesar imagen: abrir desde stream, validar formato
        file.stream.seek(0)
        try:
            img = Image.open(file.stream)
            img.verify()  # Verifica que sea una imagen válida
        except UnidentifiedImageError:
            flash('El archivo no es una imagen válida.', 'danger')
            return redirect(url_for('perfil'))
        file.stream.seek(0)
        img = Image.open(file.stream).convert('RGB')
        min_side = min(img.size)
        left = (img.width - min_side) // 2
        top = (img.height - min_side) // 2
        img = img.crop((left, top, left+min_side, top+min_side))
        img = img.resize((300, 300))
        img.save(path, format='PNG', quality=90)
        # Actualizar ruta en la base de datos directamente
        from db_config import get_connection
        conexion = get_connection()
        cursor = conexion.cursor()
        cursor.execute("UPDATE empleados SET avatar_url = %s WHERE id = %s", (f"img/{filename}", current_user.id))
        conexion.commit()
        conexion.close()
        # Recargar usuario en sesión para reflejar el avatar
        from flask_login import login_user
        from db_config import get_connection
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("SELECT * FROM empleados WHERE id = %s", (current_user.id,))
        usuario = cursor.fetchone()
        conexion.close()
        if usuario:
            user = Usuario()
            user.id = usuario['id']
            user.nombre = usuario['nombre']
            user.apellido = usuario['apellido']
            user.email = usuario['correo']
            user.rol = usuario['rol']
            user.telefono = usuario.get('telefono', '')
            user.password = usuario['password']
            user.equipo_id = usuario.get('equipo_id')
            user.avatar_url = usuario.get('avatar_url')
            login_user(user)
        flash('Avatar actualizado correctamente.', 'success')
    except Exception as e:
        flash('Error al procesar la imagen. Intenta con otra imagen válida.', 'danger')
        print('Error avatar:', e)
    return redirect(url_for('perfil'))

@app.route('/equipos/lista_ajax')
@login_required
def equipos_lista_ajax():
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT * FROM equipos ORDER BY nombre")
    equipos = cursor.fetchall()
    equipos_con_miembros = []
    for equipo in equipos:
        cursor.execute("""
            SELECT * FROM empleados WHERE equipo_id = %s AND rol IN ('lider', 'practicante')
        """, (equipo['id'],))
        miembros = cursor.fetchall()
        equipos_con_miembros.append({
            'id': equipo['id'],
            'nombre': equipo['nombre'],
            'idBoard': equipo.get('idBoard'),
            'miembros': miembros,
            'total_miembros': len(miembros)
        })
    conexion.close()
    # Renderizar solo el grid de equipos (sin el layout base)
    from flask import render_template
    return render_template('equipos_grid.html', equipos=equipos_con_miembros)

@app.route('/usuarios/editar', methods=['POST'])
def editar_usuario_api():
    id = request.form.get('usuario_id')
    nombre = request.form.get('nombre')
    apellido = request.form.get('apellido')
    correo = request.form.get('correo')
    rol = request.form.get('rol')
    # El equipo puede venir como nombre o id, aquí solo lo dejamos como ejemplo
    # Si necesitas actualizar el equipo, deberías buscar el id correspondiente
    # equipo = request.form.get('equipo')
    if not id:
        return 'ID de usuario requerido', 400
    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            UPDATE empleados SET nombre=%s, apellido=%s, correo=%s, rol=%s
            WHERE id=%s
        """, (nombre, apellido, correo, rol, id))
        conexion.commit()
        conexion.close()
        return redirect(url_for('estadisticas'))
    except Exception as e:
        return f'Error al editar usuario: {str(e)}', 500

@app.route('/usuarios/eliminar', methods=['POST'])
@login_required
def eliminar_usuario_api():
    if current_user.rol != 'jefe':
        return 'Acceso restringido solo para jefes', 403
    id = request.form.get('usuario_id')
    if not id:
        return 'ID de usuario requerido', 400
    try:
        archivar_tareas_usuario(id)
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)
        # Eliminar asignaciones
        cursor.execute("DELETE FROM asignaciones WHERE empleado_id = %s", (id,))
        # Eliminar de Trello si corresponde
        cursor.execute("SELECT equipo_id, correo FROM empleados WHERE id=%s", (id,))
        row = cursor.fetchone()
        equipo_id = row['equipo_id'] if row else None
        correo = row['correo'] if row else None
        if equipo_id and correo:
            cursor.execute("SELECT idBoard FROM equipos WHERE id=%s", (equipo_id,))
            row = cursor.fetchone()
            idBoard = row['idBoard'] if row else None
            if idBoard:
                eliminar_de_tablero_trello(idBoard, correo)
        cursor.execute("DELETE FROM empleados_trello WHERE empleado_id = %s", (id,))
        cursor.execute("DELETE FROM empleados WHERE id=%s", (id,))
        conexion.commit()
        conexion.close()
        return redirect(url_for('estadisticas'))
    except Exception as e:
        return f'Error al eliminar usuario: {str(e)}', 500

@app.route('/jefes/descender_rol_lider', methods=['POST'])
@login_required
def descender_rol_lider():
    if current_user.rol != 'jefe':
        flash('Acceso restringido solo para jefes.', 'danger')
        return redirect(url_for('jefes'))
    lider_id = request.form.get('lider_id')
    if not lider_id:
        flash('Debe seleccionar un líder.', 'danger')
        return redirect(url_for('jefes'))
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, nombre, apellido, rol FROM empleados WHERE id = %s", (lider_id,))
        usuario = cursor.fetchone()
        if not usuario or usuario['rol'] != 'lider':
            flash('El usuario seleccionado no es un líder válido.', 'danger')
            conexion.close()
            return redirect(url_for('jefes'))
        cursor.execute("UPDATE empleados SET rol = 'practicante' WHERE id = %s", (lider_id,))
        conexion.commit()
        crear_y_emitir_notificacion(lider_id, 'Has sido descendido a Practicante.', 'rol')
        flash('El líder ha sido descendido a practicante correctamente.', 'success')
    except Exception as e:
        conexion.rollback()
        flash(f'Error al descender líder: {str(e)}', 'danger')
    finally:
        conexion.close()
    return redirect(url_for('jefes'))

@app.route('/jefes/cambiar_equipo_usuario', methods=['POST'])
@login_required
def cambiar_equipo_usuario():
    if current_user.rol != 'jefe':
        flash('Acceso restringido solo para jefes.', 'danger')
        return redirect(url_for('jefes'))
    usuario_id = request.form.get('usuario_id')
    equipo_id = request.form.get('equipo_id')
    if not usuario_id or not equipo_id:
        flash('Debe seleccionar un usuario y un equipo.', 'danger')
        return redirect(url_for('jefes'))
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, nombre, apellido, rol FROM empleados WHERE id = %s", (usuario_id,))
        usuario = cursor.fetchone()
        if not usuario:
            flash('Usuario no encontrado.', 'danger')
            conexion.close()
            return redirect(url_for('jefes'))
        if usuario['rol'] == 'lider':
            cursor.execute("SELECT COUNT(*) as total FROM empleados WHERE equipo_id = %s AND rol = 'lider'", (equipo_id,))
            total_lideres = cursor.fetchone()['total']
            if total_lideres >= 2:
                flash('El equipo seleccionado ya tiene 2 líderes.', 'danger')
                conexion.close()
                return redirect(url_for('jefes'))
        cursor.execute("UPDATE empleados SET equipo_id = %s WHERE id = %s", (equipo_id, usuario_id))
        conexion.commit()
        crear_y_emitir_notificacion(usuario_id, f'Tu equipo ha sido cambiado.', 'rol')
        flash('El usuario ha sido cambiado de equipo correctamente.', 'success')
    except Exception as e:
        conexion.rollback()
        flash(f'Error al cambiar de equipo: {str(e)}', 'danger')
    finally:
        conexion.close()
    return redirect(url_for('jefes'))

# --- NUEVO: Tabla y lógica de tareas archivadas ---

def archivar_tareas_usuario(usuario_id):
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Crear directorio si no existe
    os.makedirs('archivos_archivados', exist_ok=True)
    # Obtener tareas activas del usuario
    cursor.execute("SELECT * FROM tareas_trello WHERE empleado_id = %s", (usuario_id,))
    tareas = cursor.fetchall()
    for tarea in tareas:
        archivos_guardados = []
        # Descargar archivos adjuntos de Trello si existe idCardTrello
        if tarea.get('idCardTrello'):
            url_attachments = f"https://api.trello.com/1/cards/{tarea['idCardTrello']}/attachments"
            params = {"key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
            resp = requests.get(url_attachments, params=params)
            if resp.status_code == 200:
                for adj in resp.json():
                    nombre = adj['name']
                    url_file = adj['url']
                    ruta_local = f"archivos_archivados/{tarea['id']}_{nombre}"
                    try:
                        r = requests.get(url_file, stream=True, timeout=10)
                        content_type = r.headers.get('Content-Type', '')
                        # Solo guardar si es un archivo descargable
                        if 'text/html' not in content_type and r.status_code == 200:
                            with open(ruta_local, 'wb') as f:
                                shutil.copyfileobj(r.raw, f)
                            archivos_guardados.append({'nombre': nombre, 'ruta': ruta_local})
                        else:
                            # Si es un enlace externo, guardar solo el enlace
                            archivos_guardados.append({'nombre': nombre, 'ruta': url_file})
                    except Exception as e:
                        # Si falla la descarga, guardar solo el enlace
                        archivos_guardados.append({'nombre': nombre, 'ruta': url_file})
        # Insertar en tareas_archivadas
        cursor.execute("""
            INSERT INTO tareas_archivadas (titulo, descripcion, equipo_nombre, fecha_creacion, prioridad, estado, idCardTrello)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            tarea['titulo'], tarea['descripcion'], tarea.get('equipo_id', ''), tarea['fecha_creacion'], tarea['prioridad'], tarea['estado'], tarea.get('idCardTrello')
        ))
        tarea_archivada_id = cursor.lastrowid
        # Guardar archivos en la tabla archivos_tareas_archivadas
        for archivo in archivos_guardados:
            cursor.execute("""
                INSERT INTO archivos_tareas_archivadas (tarea_archivada_id, nombre, ruta)
                VALUES (%s, %s, %s)
            """, (tarea_archivada_id, archivo['nombre'], archivo['ruta']))
        # Eliminar de Trello si corresponde
        if tarea.get('idCardTrello'):
            eliminar_tarjeta_trello(tarea['idCardTrello'])
        # Eliminar de tareas_trello
        cursor.execute("DELETE FROM tareas_trello WHERE id = %s", (tarea['id'],))
    conexion.commit()
    cursor.close()
    conexion.close()

# Endpoint para mostrar tareas archivadas (solo jefe)
@app.route('/tareas/archivadas')
@login_required
def ver_tareas_archivadas():
    if current_user.rol != 'jefe':
        return 'Acceso restringido solo para jefes', 403
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tareas_archivadas ORDER BY fecha_archivado DESC")
    tareas = cursor.fetchall()
    # Obtener todos los usuarios (líderes y practicantes)
    cursor.execute("SELECT id, nombre, apellido, rol, equipo_id FROM empleados WHERE rol IN ('lider', 'practicante') ORDER BY nombre")
    usuarios = cursor.fetchall()
    # Obtener todos los equipos
    cursor.execute("SELECT id, nombre FROM equipos ORDER BY nombre")
    equipos = cursor.fetchall()
    # Obtener archivos adjuntos de tareas archivadas
    cursor.execute("SELECT * FROM archivos_tareas_archivadas")
    archivos_tareas_archivadas = cursor.fetchall()
    conexion.close()
    return render_template('tareas_archivadas.html', tareas=tareas, usuarios=usuarios, equipos=equipos, archivos_tareas_archivadas=archivos_tareas_archivadas)

# Endpoint para reasignar tarea archivada
@app.route('/tareas/archivadas/reasignar', methods=['POST'])
@login_required
def reasignar_tarea_archivada():
    if current_user.rol != 'jefe':
        return 'Acceso restringido solo para jefes', 403
    tarea_id = request.form.get('tarea_id')
    nuevo_empleado_id = request.form.get('empleado_id')
    nuevo_equipo_id = request.form.get('equipo_id')
    if not tarea_id or not nuevo_empleado_id or not nuevo_equipo_id:
        return 'Datos incompletos', 400
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Obtener datos de la tarea archivada
    cursor.execute("SELECT * FROM tareas_archivadas WHERE id = %s", (tarea_id,))
    tarea = cursor.fetchone()
    if not tarea:
        conexion.close()
        return 'Tarea no encontrada', 404
    # Obtener idBoard del equipo
    cursor.execute('SELECT idBoard FROM equipos WHERE id = %s', (nuevo_equipo_id,))
    row = cursor.fetchone()
    idBoard = row['idBoard'] if row else None
    idCardTrello = None
    if idBoard:
        # Crear tarjeta en Trello
        idCardTrello = crear_tarjeta_trello(idBoard, tarea['titulo'], tarea['descripcion'])
    # Crear tarea activa
    cursor.execute("""
        INSERT INTO tareas_trello (titulo, descripcion, fecha_creacion, prioridad, estado, equipo_id, empleado_id, idCardTrello)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        tarea['titulo'], tarea['descripcion'], tarea['fecha_creacion'], tarea['prioridad'], tarea['estado'], nuevo_equipo_id, nuevo_empleado_id, idCardTrello
    ))
    # Eliminar de tareas_archivadas
    cursor.execute("DELETE FROM tareas_archivadas WHERE id = %s", (tarea_id,))
    conexion.commit()
    conexion.close()
    return redirect(url_for('ver_tareas_archivadas'))

# Helper para eliminar tarjeta de Trello

def eliminar_tarjeta_trello(idCardTrello):
    url = f"https://api.trello.com/1/cards/{idCardTrello}"
    params = {"key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
    requests.delete(url, params=params)
    return True

def crear_tabla_archivos_tareas_archivadas():
    conexion = get_connection()
    cursor = conexion.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS archivos_tareas_archivadas (
            id INT AUTO_INCREMENT PRIMARY KEY,
            tarea_archivada_id INT,
            nombre VARCHAR(255),
            ruta VARCHAR(255),
            FOREIGN KEY (tarea_archivada_id) REFERENCES tareas_archivadas(id) ON DELETE CASCADE
        )
    ''')
    conexion.commit()
    conexion.close()

crear_tabla_archivos_tareas_archivadas()

@app.route('/tareas/archivadas/actualizar_adjuntos/<int:tarea_id>', methods=['POST'])
@login_required
def actualizar_adjuntos_tarea_archivada(tarea_id):
    if current_user.rol != 'jefe':
        return {'success': False, 'message': 'Acceso restringido solo para jefes.'}, 403
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    # Buscar la tarea archivada y su idCardTrello original (si existe)
    cursor.execute("SELECT * FROM tareas_archivadas WHERE id = %s", (tarea_id,))
    tarea = cursor.fetchone()
    if not tarea:
        conexion.close()
        return {'success': False, 'message': 'Tarea archivada no encontrada.'}, 404
    # Buscar en tareas_trello_archivadas (si tienes una tabla de respaldo) o en tareas_trello eliminadas
    # Aquí asumimos que guardaste el id original en la columna equipo_nombre o similar
    idCardTrello = tarea.get('idCardTrello')
    if not idCardTrello:
        conexion.close()
        return {'success': False, 'message': 'No se encontró el idCardTrello original para esta tarea.'}, 404
    # Descargar archivos adjuntos de Trello
    os.makedirs('archivos_archivados', exist_ok=True)
    url_attachments = f"https://api.trello.com/1/cards/{idCardTrello}/attachments"
    params = {"key": TRELLO_API_KEY, "token": TRELLO_API_TOKEN}
    resp = requests.get(url_attachments, params=params)
    archivos_guardados = []
    if resp.status_code == 200:
        for adj in resp.json():
            nombre = adj['name']
            url_file = adj['url']
            ruta_local = f"archivos_archivados/{tarea_id}_{nombre}"
            try:
                r = requests.get(url_file, stream=True, timeout=10)
                content_type = r.headers.get('Content-Type', '')
                if 'text/html' not in content_type and r.status_code == 200:
                    with open(ruta_local, 'wb') as f:
                        shutil.copyfileobj(r.raw, f)
                    archivos_guardados.append({'nombre': nombre, 'ruta': ruta_local})
                else:
                    archivos_guardados.append({'nombre': nombre, 'ruta': url_file})
            except Exception as e:
                archivos_guardados.append({'nombre': nombre, 'ruta': url_file})
    # Guardar archivos en la tabla archivos_tareas_archivadas (evitar duplicados)
    for archivo in archivos_guardados:
        cursor.execute("SELECT COUNT(*) as total FROM archivos_tareas_archivadas WHERE tarea_archivada_id = %s AND nombre = %s", (tarea_id, archivo['nombre']))
        if cursor.fetchone()['total'] == 0:
            cursor.execute("""
                INSERT INTO archivos_tareas_archivadas (tarea_archivada_id, nombre, ruta)
                VALUES (%s, %s, %s)
            """, (tarea_id, archivo['nombre'], archivo['ruta']))
    conexion.commit()
    conexion.close()
    return {'success': True}

@app.route('/jefes/eliminar/<int:jefe_id>', methods=['POST'])
@login_required
def eliminar_jefe(jefe_id):
    if current_user.rol != 'jefe':
        return {'success': False, 'message': 'Acceso restringido solo para jefes.'}, 403
    conexion = get_connection()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT COUNT(*) as total FROM empleados WHERE rol = 'jefe'")
    total_jefes = cursor.fetchone()['total']
    if total_jefes <= 1:
        conexion.close()
        return {'success': False, 'message': 'Debe haber al menos un jefe en el sistema.'}, 400
    cursor.execute("DELETE FROM empleados WHERE id = %s AND rol = 'jefe'", (jefe_id,))
    conexion.commit()
    conexion.close()
    return {'success': True, 'message': 'Jefe eliminado correctamente.'}

if __name__ == '__main__':
    socketio.run(app, debug=True)
